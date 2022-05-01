from glob import glob
import openpyxl
import pandas as pd

import sys

# python aggregate.py src/20210112/
args = sys.argv

# 関数パート
def to_minute(cell):
    if cell.value == None:
        minute = 0
    else:
        minute = cell.value.hour*60+cell.value.minute
    
    return minute

def none_to_zero(cell):
    transportation_expense = 0 if cell.value is None else cell.value
    return transportation_expense

def sub_rest_time(working_hour):
    if working_hour >= 480:
        working_hour -= 60
    elif working_hour >= 360:
        working_hour -= 45
    else:
        pass
    
    return working_hour


def main():
    file_paths = glob(f'{args[1]}/*.xlsx')
    first_rows = [7, 13, 19, 25, 31, 37, 43]
    data = {}
    # キー：　名前、　バリュー：1週間分のデータ（datum）
    for file_path in file_paths:
        # 1週間分のデータ
        datum = {}
        datum['date'] = []
        datum['is_midnight_work'] = []
        datum['working_hour'] = []
        datum['transportation_expense'] = []


        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['週報']

        name = ws['AG4'].value

        for row in first_rows:
            date = ws.cell(row=row, column=1).value.date()

            is_midnight_work = '○' if ws.cell(row=row, column=6).value.startswith(('F', 'I', 'J', 'M')) else ''

            working_hour = to_minute(ws.cell(row=row, column=21))+to_minute(ws.cell(row=row+2, column=21))+to_minute(ws.cell(row=row+4, column=21))
            working_hour = sub_rest_time(working_hour)

            transportation_expense = none_to_zero(ws.cell(row=row, column=40))+none_to_zero(ws.cell(row=row+2, column=40))+none_to_zero(ws.cell(row=row+4, column=40))

            datum['date'].append(date)
            datum['is_midnight_work'].append(is_midnight_work)
            datum['working_hour'].append(working_hour)
            datum['transportation_expense'].append(transportation_expense)

        data[name] = datum

    wb_members = openpyxl.load_workbook('従業員集約.xlsx', data_only=True)
    for name in data.keys():
        # シートが作成済みであれば作成しない
        if not name in wb_members.sheetnames:
            wb_members.create_sheet(title=name)

    for name in data.keys():
        datum = data[name]
        _df = pd.DataFrame(datum)

        ws_members = wb_members[name]
        max_row = ws_members.max_row

        if max_row == 1:
            # カラムを設定    
            ws_members['A1'].value = '日付'
            ws_members['B1'].value = '勤務時間'
            ws_members['C1'].value = '交通費'
            ws_members['D1'].value = '深夜勤務'
            # カラム幅    
            ws_members.column_dimensions['A'].width = 15
            ws_members.column_dimensions['B'].width = 10
            ws_members.column_dimensions['C'].width = 10
            ws_members.column_dimensions['D'].width = 10

        for index, row in _df.iterrows():
            ws_members.cell(row=max_row+1+index, column=1).value = row['date']
            ws_members.cell(row=max_row+1+index, column=2).value = row['working_hour']
            ws_members.cell(row=max_row+1+index, column=3).value = row['transportation_expense']
            ws_members.cell(row=max_row+1+index, column=4).value = row['is_midnight_work']
        
        print(f'{name}さんの転記が終了')

    wb_members.save('従業員集約.xlsx')
    
if __name__=='__main__':
    main()